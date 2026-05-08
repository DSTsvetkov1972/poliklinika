import os, sys, shutil
import pandas as pd
from datetime import datetime

sys.path.append(os.getcwd())

from package.config import folders_rules_dict
from package.processors import processors_dict
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from progress.bar import FillingSquaresBar
from colorama import Fore
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl.styles.stylesheet')

def processor_starter(folder, file):
    """
    """

    folder_rules = folders_rules_dict.get(folder)
    if not folder_rules:
        return (False, f'Не созданы правила обработки для { folder }')
    
    processor_name_in_config = folder_rules.get('processor_name')
    if not processor_name_in_config:
        return (False, f'В правилах обработки не задан обработчик для { folder }')

    processor = processors_dict.get(processor_name_in_config)
    if not processor:
        return (False, f'Обработчик файлов { processor_name_in_config } не создан')
    return processor(folder, file, folders_rules_dict)





def prepared_maker():
    try:
    # if True:
        if os.path.exists(os.path.join(os.getcwd(), 'Сводка по подготовке файлов к загрузке.xlsx')):
            with open(os.path.join(os.getcwd(), 'Сводка по подготовке файлов к загрузке.xlsx'), 'r+b'):
                pass
    
        prepared_folder = os.path.join(os.getcwd(), 'Подготовленные')
        shutil.rmtree(prepared_folder)
        os.mkdir(prepared_folder)    

        processor_log = []
        no_code_dfs = []

        folders = list(os.walk('Исходники'))[0][1]

        folders = [folder for folder in folders if '_Открепление' in folder or '_Прикрепление' in folder]

        max_folder_len = max([len(f) for f in folders])

        print(Fore.BLACK)
        for folder in folders:
            files = list(os.walk(os.path.join('Исходники', folder)))[0][2]

            res_dfs = []
            
            if files:
                    
                bar = FillingSquaresBar(
                    f'{folder:>{max_folder_len}}',
                    max=len(files),
                    suffix = '%(index)d/%(max)d',
                    fill='█', empty_fill='░',
                    width = 50)    

                bar.start()

                # for file in tqdm(files, desc=folder, unit='Файл', leave=True):
                for file in files:

                    no_code_df =pd.DataFrame()
                    processor_starter_res = processor_starter(folder, file)
                    # print(file, processor_starter_res)

                    if processor_starter_res[0]:
                        res_df = processor_starter_res[1]

                        if 'Код ПИКОМЕД' in res_df.columns:
                            no_code_df = res_df[res_df['Код ПИКОМЕД']=='']
                            #print(no_code_df[['Вид медицинского обслуживания','Код ПИКОМЕД']])
                            

                        if no_code_df.empty:
                            processor_log.append({
                                'Папка': folder,
                                'Файл': file,
                                'Результат обработки': 'Ok',
                                'Строк без кода ПИКОМЕД': 0,
                                'Строк в файле': len(res_df)
                                })
                        else:
                            no_code_dfs.append(no_code_df)
                            #print(no_code_df)
                            processor_log.append({
                                'Папка': folder,
                                'Файл': file,
                                'Результат обработки': 'Не все коды сопоставлены',
                                'Строк без кода ПИКОМЕД': len(no_code_df),
                                'Строк в файле': len(res_df)
                                })
                            
                        res_dfs.append(res_df)
                    else:
                        processor_log.append({
                            'Папка': folder,
                            'Файл': file,
                            'Результат обработки': processor_starter_res[1],
                            'Строк без кода ПИКОМЕД': None,
                            'Строк в файле': None
                            })

                    if res_dfs:
                        res_df = pd.concat(res_dfs)

                        prepared_file = os.path.join(os.getcwd(), 'Подготовленные', f'{ folder }.xlsx')
                        res_df.to_excel(prepared_file, index=False)

                        wb = load_workbook(prepared_file)
                        ws = wb.active
                        
                        ws.freeze_panes = 'A2'
                        ws.auto_filter.ref = ws.dimensions 

                        for col in range(1, ws.max_column+1):
                            header_cell = ws.cell(column=col, row=1)
                            ws.column_dimensions[get_column_letter(col)].width = 24
                            if header_cell.value in ['Дата рождения', 'Период обслуживания c', 'Период обслуживания по', 'Дата открепления']:
                                for row in range(2, ws.max_row+1):
                                    cell_to_format = ws.cell(row=row, column=col)
                                    cell_to_format.number_format = 'DD.MM.YYYY'
                                    cell_to_format.font = Font(bold=True)

                        wb.save(prepared_file)

                    bar.next()

                bar.finish()
        print(Fore.RESET)

        
        no_code_res_file = os.path.join(os.getcwd(), 'Категории без кодов.xlsx')           
        if no_code_dfs:
            no_code_res_df = pd.concat(no_code_dfs)
            no_code_res_df = no_code_res_df[['Папка','Вид медицинского обслуживания']]
            no_code_res_df['Код ПИКОМЕД'] = '-'
            no_code_res_df = no_code_res_df.drop_duplicates()
            # print(no_code_res_df)

            
            no_code_res_df.to_excel(no_code_res_file, index=False)

            wb = load_workbook(no_code_res_file)
            ws = wb.active
            
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = ws.dimensions 

            # Устанавливаем ширину для конкретной колонки
            ws.column_dimensions['A'].width = 36
            ws.column_dimensions['B'].width = 72
            ws.column_dimensions['C'].width = 20

            for col in range(1, 4):
                cell = ws.cell(column=col, row=1)
                cell.font = Font(bold=True)  # Жирный шрифт
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Выравнивание по центру

            for row in range(2, ws.max_row+1):
                ws.cell(column=1, row=row).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws.cell(column=2, row=row).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                ws.cell(column=3, row=row).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            wb.save(no_code_res_file)
        else:    
            pd.DataFrame(columns=['Папка','Вид медицинского обслуживания','Код ПИКОМЕД']).to_excel(no_code_res_file, index=None)

        log_df = pd.DataFrame(processor_log)
        log_df.to_excel('Сводка по подготовке файлов к загрузке.xlsx', index=None)
            
        wb = load_workbook('Сводка по подготовке файлов к загрузке.xlsx')

        ws = wb['Sheet1']

        ws.freeze_panes = 'A2'

        # Устанавливаем ширину для конкретной колонки
        ws.column_dimensions['A'].width = 36
        ws.column_dimensions['B'].width = 42
        ws.column_dimensions['C'].width = 48
        ws.column_dimensions['D'].width = 48
        ws.column_dimensions['E'].width = 48                


        for col in range(1, 6):
            cell = ws.cell(column=col, row=1)
            cell.font = Font(bold=True)  # Жирный шрифт
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Выравнивание по центру

        ws.auto_filter.ref = ws.dimensions

        for row in range(2, ws.max_row+1):
            for column in range(3, 6):
                ws.cell(column=column, row=row).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


        wb.save('Сводка по подготовке файлов к загрузке.xlsx')

        if no_code_dfs:
            return (False, 'Есть виды медицинского обслуживания с несопоставленными кодами')
        else:
            return (True,)
            

    except Exception as e:
        return (False, e)


    

if __name__ == '__main__':
    folder = 'ВСК_Открепление'
    file = 'Согаз изменение объема пример.xls'
    processor_starter(folder, file)
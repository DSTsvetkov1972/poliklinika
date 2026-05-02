import os, sys, shutil
import pandas as pd
import zipfile
import py7zr

sys.path.append(os.getcwd())

from package.config import folders_rules_dict
from package.email_separators import separators_dict
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from package.fns import get_file_path
import re
from progress.bar import FillingSquaresBar
from colorama import Fore
from dotenv import load_dotenv

load_dotenv(os.path.join(os.getcwd(), '.config'))


def drop_files(folder, extensions):
    files = list(os.walk(os.path.join(os.getcwd(), 'Исходники', folder)))[0][2]

    for file in files:
        extension = file.split('.')[-1]
        if extension in extensions:
            os.remove(os.path.join(os.getcwd(), 'Исходники', folder, file))
    

def extract_encrypted_zip(zip_path,
                          extract_path,
                          password):

    try:
  
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            # Пароль нужно передать в байтовом формате
            zip_ref.extractall(path=extract_path, pwd=password.encode('utf-8'))

        os.remove(zip_path)
        unzipped_files =  list(os.walk(extract_path))[0][2]
        
        return(unzipped_files, f"✅ Извлечены файлы из архива {zip_path.replace(os.getcwd(), '')}")

    except RuntimeError as e:
        return (False, f"❌ {extract_path}: Неверный пароль или архив поврежден. {e}")
    except zipfile.BadZipFile:
        return(False, f"❌ {extract_path}: Файл не является ZIP-архивом или поврежден.")
    except Exception as e:
        return(False, f"❌ {extract_path}: {e}")


def extract_encrypted_7z(archive_path,
                         extract_to_folder,
                         password):
    try:
        with py7zr.SevenZipFile(archive_path, mode='r', password=password) as archive:
            archive.extractall(path=extract_to_folder)

        os.remove(archive_path)
        unzipped_files =  list(os.walk(extract_to_folder))[0][2]
        
        return(unzipped_files, f"✅ Извлечены файлы из архива {archive_path.replace(os.getcwd(), '')}")
    except Exception as e:
        return(False, f"❌ {archive_path}: {e}")   
    


def unzip_files(folder, password=''):
    source_files = list(os.walk(os.path.join(os.getcwd(), 'Исходники', folder)))[0][2]
    zip_file_paths = [os.path.join(os.getcwd(), 'Исходники', folder, file) for file in source_files if file.split(".")[-1] in ('zip', '7z')]

    for zip_file_path in zip_file_paths:
        extract_path =  os.path.join(os.getcwd(), 'Исходники', folder, "unzipped")
        if os.path.exists(extract_path):
            shutil.rmtree(extract_path)

        if zip_file_path.split('.')[-1] == 'zip':
            extract_encrypted_zip_res = extract_encrypted_zip(zip_file_path, extract_path, password)
        elif zip_file_path.split('.')[-1] == '7z':
            extract_encrypted_zip_res = extract_encrypted_7z(zip_file_path, extract_path, password)

        unzipped_files = extract_encrypted_zip_res[0]
        
        if unzipped_files:
            for unzipped_file in unzipped_files:
                unzipped_file_path = os.path.join(extract_path, unzipped_file)
                unzipped_file_path_checked = get_file_path(file_name=unzipped_file, downloaded_folder=folder)
                # unzipped_file_path_checked = file_path_if_exists_2(unzipped_file, os.path.join(os.getcwd(), 'Исходники', folder))
                shutil.move(unzipped_file_path, unzipped_file_path_checked)

            print(Fore.GREEN, extract_encrypted_zip_res[1], Fore.RESET)
        else:
            print(Fore.RED, extract_encrypted_zip_res[1], Fore.RESET)
        
        shutil.rmtree(extract_path)    


def processor_starter(folder, file):
    """
    """

    folder_rules = folders_rules_dict.get(folder)
    if not folder_rules:
        return (False, f'Не созданы правила обработки для { folder }')
    
    processor_name_in_config = folder_rules.get('separator_name')
    if not processor_name_in_config:
        return (False, f'В правилах обработки не задан обработчик для { folder }')

    processor = separators_dict.get(processor_name_in_config)
    if not processor:
        return (False, f'Обработчик файлов { processor_name_in_config } не создан')
    return processor(folder, file, folders_rules_dict)


def separator():
    unzip_files(folder='ЗЕТТА_Скачано', password=os.getenv("ZETTA_PASSWORD"))
    
    unzip_files(folder='Росгосстрах ТЭК_Скачано', password=os.getenv("RGS_TEK_PASSWORD"))

    unzip_files(folder='Росгосстрах_Скачано', password=os.getenv("RGS_PASSWORD"))
    
    unzip_files(folder='Совкомбанк_Скачано')
    drop_files(folder='Совкомбанк_Скачано', extensions = ['png'])


    try:
        processor_log = []

        folders = list(os.walk('Исходники'))[0][1]
        downloaded_folders = [folder for folder in folders if '_Скачано' in folder]
        
        max_folder_len = max([len(f) for f in downloaded_folders])

        print(Fore.BLACK)
        for folder in downloaded_folders:
            #print(Fore.MAGENTA, folder, Fore.RESET)
            files = list(os.walk(os.path.join('Исходники', folder)))[0][2]

                        
            if files:

                bar = FillingSquaresBar(
                    f'{folder:>{max_folder_len}}',
                    max=len(files),
                    suffix = '%(index)d/%(max)d',
                    fill='█', empty_fill='░',
                    width = 50)    
                
                bar.start()

                for file in files:
                    processor_starter_res = processor_starter(folder, file)
                    processor_log.append([folder, file, processor_starter_res[1]])

                    bar.next()
                
                bar.finish()

        print(Fore.RESET)                

        log_df = pd.DataFrame(processor_log, columns=['Папка', 'Файл', 'Результат обработки'])
        log_df.to_excel('Сводка по распределению файлов.xlsx', index=None)
            
        wb = load_workbook('Сводка по распределению файлов.xlsx')

        ws = wb['Sheet1']

        ws.freeze_panes = 'A2'

        # Устанавливаем ширину для конкретной колонки
        ws.column_dimensions['A'].width = 36
        ws.column_dimensions['B'].width = 84
        ws.column_dimensions['C'].width = 64


        for col in range(1, 4):
            cell = ws.cell(column=col, row=1)
            cell.font = Font(bold=True)  # Жирный шрифт
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Выравнивание по центру

        #for row in range(2, ws.max_row+1):
        #    ws.cell(column=2, row=row).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        #    ws.cell(column=3, row=row).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        #    ws.cell(column=4, row=row).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.auto_filter.ref = ws.dimensions
        wb.save('Сводка по распределению файлов.xlsx')
        return (True,)

    except Exception as e:
        return (False, e)


    

if __name__ == '__main__':
    unzip_files(folder='ЗЕТТА_Скачано', password=("ZETTA_PASSWORD"))
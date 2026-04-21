import os, sys
import pandas as pd
from progress.bar import FillingSquaresBar

sys.path.append(os.getcwd())

from colorama import Fore

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from imap_tools import MailBox, AND
from package.config import IMAP_SERVER, IMAP_PORT, EMAIL, APP_PASSWORD
from package.config import folders_rules_dict


def get_unread_messages(email_folder):
    #"""Получает количество непрочитанных писем в папке"""
    #try:
        with MailBox(IMAP_SERVER, port=IMAP_PORT).login(EMAIL, APP_PASSWORD, initial_folder='INBOX') as mailbox:

                mailbox.folder.set(email_folder)
                
                # Ищем непрочитанные письма (критерий seen=False)
                unread_messages = list(mailbox.fetch(AND(seen=False), mark_seen=False))
                
                return (True, len(unread_messages))
            
    #except socket.gaierror as e:
    #    return (True, f"get_unread_messages_from_all_folders: { e }.\nВозможно отсутствует подключение к интернет.")
    #except Exception as e:
    #    return (False, f"get_unread_messages_from_all_folders: { e }")


def email_summary():
    
    email_folders = {}
    
    for folder, folder_rules  in folders_rules_dict.items():

        email_folder = folder_rules.get('email_folder')
        if email_folder:
            email_folders[folder.split('_')[0]] = email_folder

       
    
    email_folders_info = []
    
    print(Fore.BLACK)
    bar = FillingSquaresBar(
        'Просматриваем папки в электронной почте:',
        max=len(email_folders),
        suffix = '%(index)d/%(max)d',
        fill='█', empty_fill='░',
        width = 50) 
    

    for ensurence_company, email_folder in email_folders.items():
        #if ensurence_company!='Югория':
        #    continue
        #email_folders[ensurence_company] = get_unread_messages(email_folder)[1]
        email_folders_info.append({'Компания': ensurence_company, 'Непрочитанных писем': get_unread_messages(email_folder)[1]}) 
        bar.next()
        
    bar.finish()
    print(Fore.RESET)
    
    return pd.DataFrame(email_folders_info)

    
def folders_summary():
    """Получаем датафрейм
    с количеством файлов в исходных папках
    """

    downloaded_folders_info = []
    start_folders_info = []
    start_folders_2_info = []
    finish_folders_info = []

    for folder in folders_rules_dict.keys():
        folder_parts = folder.split('_')
        ensurence_company = folder_parts[0]
        file_type = '_'.join(folder_parts[1:])

        files = list(os.walk(os.path.join(os.getcwd(), 'Исходники', folder)))[0][2]
        folder_dict = {'Компания': ensurence_company, file_type: len(files)}

        if file_type == 'Скачано' in folder:
            downloaded_folders_info.append(folder_dict)
        elif file_type == 'Прикрепление' in folder:
            start_folders_info.append(folder_dict)
        elif file_type == 'Прикрепление_2' in folder:
            start_folders_2_info.append(folder_dict)            
        elif file_type == 'Открепление' in folder:
            finish_folders_info.append(folder_dict)

    downloaded_df = pd.DataFrame(downloaded_folders_info)
    start_df = pd.DataFrame(start_folders_info)
    start_2_df  = pd.DataFrame(start_folders_2_info)    
    finish_df = pd.DataFrame(finish_folders_info)

    res_df = pd.merge(downloaded_df, start_df, on='Компания', how='outer')
    res_df = pd.merge(res_df, start_2_df, on='Компания', how='outer')    
    res_df = pd.merge(res_df, finish_df, on='Компания', how='outer')

    res_df.fillna(0, inplace=True)

    return res_df 
    

def prepared_summary():
    """Получаем датафрейм
    с количеством файлов собранных в файл
    подготвленных к загрузке и количество строк в этих файлах
    """

    prepared_files = list(os.walk(os.path.join(os.getcwd(), 'Подготовленные')))[0][2]

    if not prepared_files:
         return pd.DataFrame(columns=['Компания'])

    prepared_files_info = []

    for file in prepared_files:
        df = pd.read_excel(os.path.join(os.getcwd(), 'Подготовленные', file))
        df = df[['Папка', 'Файл']]
        #df = df.groupby('Папка').agg({
        #    'Файл': 'nunique' ,
        #    'Папка': 'count'
        #    }).rename(columns={
        #       'Файл': 'Файлов в подготовленном файле',
        #       'Папка': 'Строк в подготовленном файле'
        #       })
        
        df = df.groupby('Папка').agg({
            'Файл': 'nunique'
            }).rename(columns={
               'Файл': 'Файлов в подготовленном файле'
               })
        # print(df)
        prepared_files_info.append(df)

    if prepared_files_info:
        total_df = pd.concat(prepared_files_info)
        total_df = total_df.reset_index('Папка')
        total_df['Компания'] = total_df['Папка'].apply(lambda x: x.split('_')[0])
        total_df['Тип файла'] = total_df['Папка'].apply(lambda x: '_'.join(x.split('_')[1:]))        

        total_df = total_df.pivot(index='Компания', columns='Тип файла', values='Файлов в подготовленном файле')
        total_df = total_df.fillna(0)
        total_df = total_df.reset_index()

        return total_df


def summary():
    

        email_summary_df = email_summary()
        #print(email_summary_df)
        folders_summary_df = folders_summary()
        #print(folders_summary_df)
        prepared_summary_df = prepared_summary()
        #print(prepared_summary_df)

        res_df = pd.merge(email_summary_df, folders_summary_df, on='Компания', how='outer')
        res_df = pd.merge(res_df, prepared_summary_df, on='Компания', how='outer')
        res_df = res_df.fillna(0)
        res_df = res_df.rename(
             columns = {
                  'Прикрепление_x': 'Прикрепление',
                  'Прикрепление_2_x': 'Прикрепление_2',
                  'Открепление_x': 'Открепление',
                  'Открепление_y': 'Открепление\nподготовленно',
                  'Прикрепление_y': 'Прикрепление\nподготовленно',
                  'Прикрепление_2_y': 'Прикрепление_2\nподготовленно'
             }
        )
        #print(res_df)

        res_df.to_excel('Исходники и подготовленные.xlsx', index=None)

        wb = load_workbook('Исходники и подготовленные.xlsx')

        ws = wb['Sheet1']

        ws.freeze_panes = 'A2'

        # Устанавливаем ширину для конкретной колонки
        ws.column_dimensions['A'].width = 36
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 20

        for col in range(1, 10):
            cell = ws.cell(column=col, row=1)
            cell.font = Font(bold=True)  # Жирный шрифт
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Выравнивание по центру

        for row in range(2, ws.max_row+1):
            for column in range(2, 10):
                ws.cell(column=column, row=row).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


        wb.save('Исходники и подготовленные.xlsx')



if __name__ == '__main__':
    #print(email_summary())

    #print(folders_summary())
    #print(prepared_summary())
    summary()
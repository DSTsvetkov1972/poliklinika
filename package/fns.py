import pyperclip
import os, re
from colorama import Fore
import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


def open_file():
    path = None
    try:
        clipboard = pyperclip.paste()
        #clipboard = 'РЕСО_Прикрепление_2	p41899714.xlsx'
        clipboard_parts = clipboard.replace('\r\n', '').split('\t')
        folder = clipboard_parts[0]
        folder_parts = folder.split('_')

        file = clipboard_parts[1]

        for folder_1 in ['Открепление', 'Прикрепление', 'Прикрепление_2', 'Скачано', 'Изменение']:
            supposed_folder = f'{folder_parts[0]}_{folder_1}'
            path = os.path.join(os.getcwd(), 'Исходники', supposed_folder, file)
            if os.path.exists(path):
                print(Fore.GREEN + f'Открыли:\n{file} из папки {supposed_folder}' + Fore.RESET)
                os.startfile(path)
                return
            else:
                continue
        print(Fore.RED + f'Что-то не то в буфере обмена:\n{clipboard[:150]}...' + Fore.RESET)
    except Exception as e:
        print(Fore.RED + f'Ошибка выполнения:\n{e}\nЧто-то не то в буфере обмена:\n{clipboard[:150]}...' + Fore.RESET)


def get_file_path(file_name, downloaded_folder):
    pattern_capture_1 = r'(.*)\[(\d+)\]\.(\w+)$'
    pattern_capture_2 = r'(.*)\.(\w+)$'

    
    company = downloaded_folder.split('_')[0]

    source_folders = list(os.walk(os.path.join(os.getcwd(), 'Исходники')))[0][1]
    
    company_files = []
    for folder in source_folders:
        if company in folder:
            files = list(os.walk(os.path.join(os.getcwd(), 'Исходники', folder)))[0][2]
            company_files += files

    new_file_name = file_name
    
    while True:
        if new_file_name in company_files:
            match_1 = re.match(pattern_capture_1, new_file_name)
            match_2 = re.match(pattern_capture_2, new_file_name)

            if match_1:
                path = match_1.group(1)
                number = match_1.group(2)
                extension = match_1.group(3)
                new_file_name = f"{path}[{ int(number)+1 }].{extension}"
 
            elif match_2:
                path = match_2.group(1)
                extension = match_2.group(2)
                new_file_name = f"{path}[0].{extension}"
            else:
                raise ValueError('имя файла не соответствует шаблону')
            continue
        else:
            new_file_path = os.path.join(os.getcwd(), 'Исходники', downloaded_folder, new_file_name)
            return new_file_path               


def get_code_by_category(folder, category):
    with sqlite3.connect(os.path.join(os.getcwd(), 'project.db')) as conn:
        sql = f"SELECT code FROM folder_category_code WHERE folder='{folder}' AND category='{category}'"
        cur = conn.cursor()
        cur.execute(sql)
        res = cur.fetchone()
    if res:
        return res[0]
    else:
        return ''


def add_codes():
    try:
    # if True:
        df = pd.read_excel(os.path.join(os.getcwd(), 'Категории и коды.xlsx'), dtype=str)

        for row in df.itertuples():
            folder = row[1]
            category = row[2]
            code = row[3]

            if code in [None, '',  '-']:
                return(False, f'Папка: "{folder}", категория: "{category}". Не задан код в файле "Категории без кода.xlsx"!')
            else:
                with sqlite3.connect(os.path.join(os.getcwd(), 'project.db')) as conn:
                    cur = conn.cursor()
                    cur.execute(f"DELETE FROM folder_category_code WHERE folder='{folder}' AND category='{category}'")
                    cur.execute(f"INSERT OR REPLACE INTO folder_category_code (folder, category, code) VALUES ('{folder}', '{category}', '{code}')")
            
            
            print(Fore.GREEN, f'Папка: {folder}, категория: {category}. Сопоставили код {code}', Fore.RESET)

        with sqlite3.connect(os.path.join(os.getcwd(), 'project.db')) as conn:
            cur = conn.cursor()

            cur.execute("DROP TABLE IF EXISTS folder_category_code_tmp")
            cur.execute("CREATE TABLE folder_category_code_tmp AS SELECT * FROM folder_category_code ORDER BY folder, category")
            cur.execute('DROP TABLE IF EXISTS folder_category_code')
            cur.execute('ALTER TABLE folder_category_code_tmp RENAME TO folder_category_code')

        return (True, "Загрузка кодов завершена")
    except Exception as e:
        return(False, repr(e))

   
def format_folder_category_code_table(no_code_res_file=os.path.join(os.getcwd(), 'Категории и коды.xlsx')):
            wb = load_workbook(no_code_res_file)
            ws = wb.active
            
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = ws.dimensions 

            # Устанавливаем ширину для конкретной колонки
            ws.column_dimensions['A'].width = 36
            ws.column_dimensions['B'].width = 72
            ws.column_dimensions['C'].width = 20

            for col in range(1, 4):
                cell = ws.cell(column=col, row=1)
                cell.font = Font(bold=True)  # Жирный шрифт
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Выравнивание по центру

            for row in range(2, ws.max_row+1):
                ws.cell(column=1, row=row).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws.cell(column=2, row=row).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                ws.cell(column=3, row=row).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            wb.save(no_code_res_file)    


def download_codes():
    try:
    # if True:
        with sqlite3.connect(os.path.join(os.getcwd(), 'project.db')) as conn:
            df = pd.read_sql_query("SELECT * FROM folder_category_code ORDER BY folder, category", conn)
            df.columns = ['Папка','Вид медицинского обслуживания', 'Код ПИКОМЕД']
            df.to_excel(os.path.join(os.getcwd(), 'Категории и коды.xlsx'), index=None)
            format_folder_category_code_table()
        return (True, 'Коды выгружены в таблицу "Категории и коды"')
    except Exception as e:
        return(False, repr(e))
 
    
if __name__ == '__main__':

    
    print(add_codes())
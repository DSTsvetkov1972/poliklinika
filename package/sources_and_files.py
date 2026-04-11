import os, sys
import pandas as pd

sys.path.append(os.getcwd())

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


def source_folders_summary():
    """Получаем датафрейм
    с количеством файлов в исходных папках
    """

    folders_info = []
    folders = list(os.walk(os.path.join(os.getcwd(), 'Исходники')))[0][1]
    
    for folder in folders:
        files = list(os.walk(os.path.join(os.getcwd(), 'Исходники', folder)))[0][2]
        folders_info.append([folder, len(files)])

    df = pd.DataFrame(folders_info, columns=['Папка', 'Файлов в папке'])
    return df

    

def prepared_summary():
    """Получаем датафрейм
    с количеством файлов собранных в файл
    подготвленных к загрузке и количество строк в этих файлах
    """

    prepared_files = list(os.walk(os.path.join(os.getcwd(), 'Подготовленные')))[0][2]
    prepared_files_info = []

    for file in prepared_files:
        df = pd.read_excel(os.path.join(os.getcwd(), 'Подготовленные', file))
        df = df[['Папка', 'Файл']]
        df = df.groupby('Папка').agg({
            'Файл': 'nunique',
            'Папка': 'count'
            }).rename(columns={
               'Файл': 'Файлов в подготовленном файле',
               'Папка': 'Строк в подготовленном файле'
               })
        # print(df)
        prepared_files_info.append(df)

    if prepared_files_info:
        total_df = pd.concat(prepared_files_info)
        total_df = total_df.reset_index('Папка')
    else:
        total_df = pd.DataFrame(columns=['Папка'])
    #total_df_grouped.columns = ['Папка', 'Подготовленных файлов']
    return total_df


def sources_and_prepared_summary():
    
    if os.path.exists(os.path.join(os.getcwd(),'~$Исходники и подготовленные.xlsx')):
        return False
    else:
        source_folders_summary_df = source_folders_summary()
        prepared_summary_df = prepared_summary()

        res_df = pd.merge(source_folders_summary_df, prepared_summary_df, on='Папка', how='outer')
        res_df.to_excel('Исходники и подготовленные.xlsx', index=None)

        wb = load_workbook('Исходники и подготовленные.xlsx')

        ws = wb['Sheet1']

        ws.freeze_panes = 'A2'

        # Устанавливаем ширину для конкретной колонки
        ws.column_dimensions['A'].width = 36
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

        for col in range(1, 5):
            cell = ws.cell(column=col, row=1)
            cell.font = Font(bold=True)  # Жирный шрифт
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Выравнивание по центру

        for row in range(2, ws.max_row+1):
            ws.cell(column=2, row=row).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(column=3, row=row).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(column=4, row=row).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        wb.save('Исходники и подготовленные.xlsx')
        return True



if __name__ == '__main__':

    sources_and_prepared_summary()
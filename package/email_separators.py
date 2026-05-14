import pandas as pd
import os, sys, shutil
sys.path.append(os.getcwd())
import re
import msoffcrypto
import io

from datetime import datetime
from package.config import folders_rules_dict

from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

from colorama import Fore
from openpyxl import load_workbook

#from dotenv import load_dotenv
#if not load_dotenv(os.path.join(os.getcwd(), '.config')):
#    print(Fore.RED, 'Файл конфигурации .config отсутствует в папке с ffic.exe', Fore.RESET)

    

def email_by_file_name(folder, file, folders_rules_dict):
    try:
        file_path = os.path.join(os.getcwd(), 'Исходники', folder, file)
        # print(folders_rules_dict[folder]['file_actions'])
        
        for file_rule in folders_rules_dict[folder]['file_rules']:
            #print(file_rule['pattern'], file, file_rule['pattern'] in file, re.search(file_rule['pattern'], file), sep = '\n')
            if re.search(file_rule['pattern'], file):
                if file_rule['target_folder'] == 'удалён':
                    # Удаляем файл
                    return (True, file_rule['target_folder'])
                elif file_rule['target_folder'] :
                    new_file_path = os.path.join(os.getcwd(), 'Исходники', file_rule['target_folder'], file)
                    shutil.move(file_path, new_file_path)
                    return (True, file_rule['target_folder'])
        return (True, 'Не установлены правила обработки файла')
            
    except Exception as e:
        return(False, e)


def email_by_cell_value_bkp(folder, file, folders_rules_dict):
    if file[-4:] not in ['.xls', 'xlsx']:
        return (True, 'Не установлены правила обработки файла')
    
    try:
        file_path = os.path.join(os.getcwd(), 'Исходники', folder, file)
        wb = load_workbook(file_path, data_only=True)
        
        for file_rule in folders_rules_dict[folder]['file_rules']:
            if file_rule['sheet_name'] not in wb.sheetnames:
                continue
            
            else:
                ws = wb[file_rule['sheet_name']]

                if re.search(file_rule['pattern'], str(ws[file_rule['cell']].value)):
                    new_file_path = os.path.join(os.getcwd(), 'Исходники', file_rule['target_folder'], file)
                    shutil.move(file_path, new_file_path)
                    return (True, file_rule['target_folder'])
        
        return (True, 'Не установлены правила обработки файла')    

    except Exception as e:
        return(False, e)


def email_by_cell_value(folder, file, folders_rules_dict):
    
    file_extension = file.split('.')[-1]
    
    if file_extension in folders_rules_dict[folder]['file_to_del_extensions']:
        return(True, "удалён")
    elif file_extension not in ['xls', 'xlsx']:
        return (True, 'Не установлены правила обработки файла')
    
    try:
    #if True:
        file_path = os.path.join(os.getcwd(), 'Исходники', folder, file)
        
        with pd.ExcelFile(file_path) as wb:
            wb_sheet_names = wb.sheet_names

        for file_rule in folders_rules_dict[folder]['file_rules']:
            sheet_name = file_rule['sheet_name']
            
            if sheet_name:
                if not sheet_name in wb_sheet_names:
                    continue



                df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str, header=None,  engine='calamine')
            else:
                df = pd.read_excel(file_path, dtype=str, header=None,  engine='calamine')   
                
            col_letter, row_num = coordinate_from_string(file_rule['cell'])
            col_num = column_index_from_string(col_letter)       

            cell_value = df.iat[row_num-1, col_num-1]
            # print(cell_value)
            
            if re.search(file_rule['pattern'], str(cell_value)):
                new_file_path = os.path.join(os.getcwd(), 'Исходники', file_rule['target_folder'], file)
                shutil.move(file_path, new_file_path)
                return (True, file_rule['target_folder'])
        
        return (True, 'Не установлены правила обработки файла')    

    except Exception as e:
        return(False, e)
    

def look_insight_rgs_tek_file(
        file_path = os.path.join(
            os.getcwd(),
            "Исходники",
            "Росгосстрах ТЭК_Скачано",
            '25.02.21 1 Ю761(010-1).xlsx')):

    password = os.getenv('RGS_TEK_PASSWORD')

    # 1. Открываем encrypted-файл и дешифруем его в объект BytesIO
    decrypted = io.BytesIO()
    with open(file_path, 'rb') as f:
        office_file = msoffcrypto.OfficeFile(f)
        if office_file.is_encrypted():
            office_file.load_key(password=password)  # Применяем пароль
            office_file.decrypt(decrypted)           # Расшифровываем в память

            # 2. Переводим "курсор" в начало потока и читаем файл через pandas
            decrypted.seek(0)
            df = pd.read_excel(decrypted, header=None) # engine указывать необязательно
        else:
            df = pd.read_excel(file_path, header=None, dtype=str, engine='calamine')

    # 3. Готово!
    print(df)



def email_rgs_tek(folder, file, folders_rules_dict):

    if file.split('.')[-1] in ('png'):
        return(True, "удалён")

    try:
    # if True:
        file_path = os.path.join(os.getcwd(), 'Исходники', folder, file)
        

        # 1. Открываем encrypted-файл и дешифруем его в объект BytesIO
        decrypted = io.BytesIO()
        
        with open(file_path, 'rb') as f:
            office_file = msoffcrypto.OfficeFile(f)
            if office_file.is_encrypted():
                office_file.load_key(password=os.getenv('RGS_TEK_PASSWORD'))  # Применяем пароль
                office_file.decrypt(decrypted)           # Расшифровываем в память

                # 2. Переводим "курсор" в начало потока и читаем файл через pandas
                decrypted.seek(0)
                df = pd.read_excel(decrypted, header=None, dtype=str)
            else:
                df = pd.read_excel(file_path, header=None, dtype=str, engine='calamine')

        for file_rule in folders_rules_dict[folder]['file_rules']:
            try:
                if re.search(file_rule["pattern"], df.iloc[file_rule["row"]].loc[file_rule["column"]]):
                    new_file_path = os.path.join(os.getcwd(), 'Исходники', file_rule['target_folder'], file)
                    shutil.move(file_path, new_file_path)
                    return (True, file_rule['target_folder'])
            except IndexError:
                continue
            except TypeError:
                continue
        
        return (True, 'Не установлены правила обработки файла')    
    #except msoffcrypto.exceptions.FileFormatError:
    #    return(True, "удалён")
    except Exception as e:
        return(False, repr(e))



separators_dict = {
    "email_by_file_name": email_by_file_name,
    "email_by_cell_value": email_by_cell_value,
    "email_rgs_tek": email_rgs_tek
}

if __name__ == '__main__':

    from dotenv import load_dotenv
    if not load_dotenv(os.path.join(os.getcwd(), '.config')):
        print(Fore.RED + 'Файл конфигурации .config отсутствует в папке с ffic.exe' + Fore.RESET)
        while True:
            pass

    
    folder, file = 'Росгосстрах_Скачано', '14-05 изм 010-1.xls'

    look_insight_rgs_tek_file(os.path.join(os.getcwd(), 'Исходники', folder, file))        


    
    # email_by_cell_value(folder, file, folders_rules_dict)

